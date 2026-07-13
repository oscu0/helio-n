from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from Library.SW.Stats import build_eval_index, build_score_frames


REPORT_COLUMNS = [
    (
        "ch_relative_area",
        ("input", "ch_area", "ch_relative_area"),
        "mean",
        "Средняя за час относительная площадь корональной дыры во входных данных модели. Время относится к исходному времени у Солнца, поэтому эту величину нельзя строка-в-строку считать входом прогноза у КА с той же временной меткой. Контекстная величина, в статистике качества не используется.",
        "0.000000",
    ),
    (
        "v_empirical",
        ("input", "model_input", "v_empirical"),
        "mean",
        "Средняя за час эмпирическая скорость солнечного ветра у Солнца, км/с. Время относится к запуску потока у Солнца, поэтому эту величину нельзя строка-в-строку считать входом прогноза у КА с той же временной меткой. Контекстная величина, в статистике качества не используется.",
        "0.0",
    ),
    (
        "ace_v_real",
        ("analysis", "ace_earth", "v_real"),
        "mean",
        "Средняя за час наблюдаемая скорость солнечного ветра ACE у Земли, км/с. Пропуски исходных наблюдений оставлены пустыми.",
        "0.0",
    ),
    (
        "ace_v_predict_raw",
        ("analysis", "ace_earth", "v_predict_raw"),
        "mean",
        "Средний за час сырой прогноз модели для ACE/Земли без постобработки медленного солнечного ветра, км/с.",
        "0.0",
    ),
    (
        "ace_v_predict_slow",
        ("analysis", "ace_earth", "v_predict"),
        "mean",
        "Средний за час прогноз модели для ACE/Земли с постобработкой медленного солнечного ветра, км/с. Эта серия сравнивается с SWX.",
        "0.0",
    ),
    (
        "ace_slow_sw_patch_mask",
        ("analysis", "ace_earth", "slow_sw_patch_mask"),
        "max",
        "TRUE, если в этом часовом интервале к прогнозу ACE применена постобработка медленного солнечного ветра.",
        "General",
    ),
    (
        "ace_v_1cr_ago",
        ("analysis", "ace_earth", "v_1cr_ago"),
        "mean",
        "Рекуррентный прогноз ACE: наблюдаемая скорость за один оборот Каррингтона до текущего времени, км/с. Пропуски не интерполированы.",
        "0.0",
    ),
    (
        "ace_v_swx",
        ("analysis", "ace_earth", "v_swx"),
        "mean",
        "Средний за час прогноз SWX для ACE/Земли, км/с, с пропусками в том виде, в котором они сохранены в reproduction parquet.",
        "0.0",
    ),
    (
        "ace_is_icme",
        ("analysis", "ace_earth", "is_icme"),
        "max",
        "TRUE внутри интервала ICME у Земли и 12 часов после его конца. Та же полуоткрытая маска используется в статистике статьи.",
        "General",
    ),
    (
        "stereo_a_v_real",
        ("analysis", "stereo_a", "v_real"),
        "mean",
        "Средняя за час наблюдаемая скорость солнечного ветра STEREO-A, км/с. Пропуски исходных наблюдений оставлены пустыми.",
        "0.0",
    ),
    (
        "stereo_a_v_predict_raw",
        ("analysis", "stereo_a", "v_predict_raw"),
        "mean",
        "Средний за час сырой прогноз модели для STEREO-A, км/с. Модель медленного солнечного ветра к STEREO-A не применяется.",
        "0.0",
    ),
    (
        "stereo_a_v_1cr_ago",
        ("analysis", "stereo_a", "v_1cr_ago"),
        "mean",
        "Рекуррентный прогноз STEREO-A: наблюдаемая скорость за один оборот Каррингтона до текущего времени, км/с. Пропуски не интерполированы.",
        "0.0",
    ),
    (
        "stereo_a_is_icme",
        ("analysis", "stereo_a", "is_icme"),
        "max",
        "TRUE внутри интервала ICME у STEREO-A и 12 часов после его конца. Та же полуоткрытая маска используется в статистике статьи.",
        "General",
    ),
]


def build_hourly_report_frame(
    reproduction_frame,
    comparison_frames,
    start_dt,
    end_dt,
    freq="1h",
):
    report_index = build_eval_index(start_dt=start_dt, end_dt=end_dt, freq=freq)
    run_frame = reproduction_frame.copy()
    run_frame.index = pd.to_datetime(run_frame.index)
    run_frame = run_frame.loc[
        (run_frame.index >= pd.Timestamp(start_dt))
        & (run_frame.index < pd.Timestamp(end_dt))
    ]
    assert not run_frame.empty, (
        f"No rows in exact run window [{start_dt}, {end_dt}) after loading "
        "the matching reproduction parquet"
    )

    required_input_columns = [
        source_column
        for _name, source_column, _agg, _description, _number_format in REPORT_COLUMNS
        if source_column[0] == "input"
    ]
    missing_input_columns = [
        source_column
        for source_column in required_input_columns
        if source_column not in run_frame.columns
    ]
    assert not missing_input_columns, (
        f"Missing reproduction input columns: {missing_input_columns}"
    )

    score_frames = build_score_frames(
        comparison_frames=comparison_frames,
        start_dt=start_dt,
        end_dt=end_dt,
        freq=freq,
    )
    report_frame = pd.DataFrame(index=report_index)
    for name, source_column, aggregation, _description, _number_format in REPORT_COLUMNS:
        source_kind = source_column[0]
        if source_kind == "input":
            report_frame[name] = (
                run_frame[source_column]
                .resample(freq)
                .agg(aggregation)
                .reindex(report_index)
            )
        else:
            assert source_kind == "analysis"
            _source_kind, sat_name, column = source_column
            report_frame[name] = score_frames[sat_name][column].reindex(report_index)

    report_frame.insert(0, "time", report_frame.index)
    return report_frame


def write_report(report_frame, output_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    header_specs = [
        (
            "time",
            "Начало часового интервала, UTC. Диапазон отчета полуоткрытый: начало включено, конец не включен.",
            "yyyy-mm-dd hh:mm",
        )
    ] + [
        (name, description, number_format)
        for name, _source_column, _aggregation, description, number_format in REPORT_COLUMNS
    ]

    for column_index, (name, description, _number_format) in enumerate(
        header_specs, start=1
    ):
        cell = worksheet.cell(row=1, column=column_index, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.comment = Comment(description, "Codex")

    for row_index, row in enumerate(
        report_frame.itertuples(index=False, name=None), start=2
    ):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            elif hasattr(value, "item"):
                value = value.item()
            worksheet.cell(row=row_index, column=column_index, value=value)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(header_specs))}{len(report_frame) + 1}"
    )
    worksheet.row_dimensions[1].height = 38

    for column_index, (name, _description, number_format) in enumerate(
        header_specs, start=1
    ):
        column_letter = get_column_letter(column_index)
        if name == "time":
            width = 20
        elif name.endswith("_mask") or name.endswith("_is_icme"):
            width = 23
        else:
            width = min(max(len(name) + 2, 16), 24)
        worksheet.column_dimensions[column_letter].width = width

        for cell in worksheet[column_letter][1:]:
            cell.number_format = number_format

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    loaded = load_workbook(output_path, read_only=False, data_only=True)
    loaded_sheet = loaded["Data"]
    expected_headers = [name for name, _description, _format in header_specs]
    actual_headers = [
        loaded_sheet.cell(row=1, column=column_index).value
        for column_index in range(1, len(header_specs) + 1)
    ]
    assert actual_headers == expected_headers
    assert loaded_sheet.max_row == len(report_frame) + 1
    assert loaded_sheet.max_column == len(header_specs)
    assert loaded_sheet["A2"].value == report_frame.iloc[0]["time"].to_pydatetime()
    assert loaded_sheet["A1"].comment.text.startswith("Начало часового интервала")
    assert "строка-в-строку" in loaded_sheet["B1"].comment.text
    assert "строка-в-строку" in loaded_sheet["C1"].comment.text
    loaded.close()
    return output_path
